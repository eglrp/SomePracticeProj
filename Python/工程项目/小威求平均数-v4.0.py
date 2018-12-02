#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
# Author: Neverland_LY
# Date: 2018-11-26
# Version: v4.0
#
# Description:
#       1. This code is used fo calculate the average of the all data filds' average except the column of 'I'
#       which represent the wind direction and the column of 'L' which represent the Hi direction.
#       2. This code will ignore the first two lines of the xlsx-file because of its own reasons.
#
# Usage:
#       1. Modify the input_file content of your own path.
#       2. Modify the output_name of your own (optional).
#       3. Think twice before changing the base_span.
#       4. You have to update the xlsx-file's row when changing the ignore_line.
#       5. Any problems or suggests, contacting with the author please.

# Customization
input_file = "E:\\xiaowei\\123.xlsx" # the file to be processed
output_name = "Result_LY.xlsx" # output name
base_span = 60 # means calculated an average number every 60 lines
ignore_line = 2 # ignore the first 2 lines of xlsx file

#####################################################################################
#                                                                                   #
#                 Do not change the following codes if unnecessary!                 #
#                                                                                   #
#####################################################################################
import openpyxl
import os, sys, winsound, datetime


# Calculate the all data filds' average
def calculateAllDataFildaverage(row_num, col_num, span = 60, ignore = 0):
    # define output file
    result_wb = openpyxl.Workbook()
    result_sheet = result_wb.active
    # define some important numbers
    block_index = 0 # a block_index means the valid row in a range of base_span
    flag_data_time = False # used to smple the datetime
    day_num = 1 # used to record the day
    start_row_num = 0 + ignore + 1 # used to record the begin_line before every loop
    current_hour_in_one_day = 1 # used to fill the second column
    global sheet # global variable
    while block_index < ( 24 * 365): # how many hous in one year (2017 is ordinary year)
        if start_row_num >= row_num:
            break
        span_length = getValidLength(start_row_num, span, day_num) # calc the true span
        date = "1997/1/1" # initial date (modifiable)
        for col_index in range(col_num): # from column A to column AJ, 36 columns totally
            if flag_data_time is False: # just used to fill first two columns
                if col_index == 0:
                    date = sheet.cell(row = block_index * span + ignore + 1, column = col_index + 1).value
                    result_sheet.cell(row = block_index + 1, column = 1).value = date.date()
                    continue
                if col_index == 1:
                    if current_hour_in_one_day > 24:
                        current_hour_in_one_day = current_hour_in_one_day - 24
                    # you can set the content of the cell which being filled
                    result_sheet.cell(row = block_index + 1, column = 2).value = str(current_hour_in_one_day)
                    flag_data_time = True
                    continue
            if col_index == 8 or col_index == 11: # ignore the column of I and L
                result_sheet.cell(row = block_index + 1, column = col_index + 1).value = "---"
                continue
            # kernal code of calculating the average
            result_sheet.cell(row = block_index + 1, column = col_index + 1).value =\
                    averageCol(sheet, start_row_num, span_length, col_index + 1)
        current_hour_in_one_day = current_hour_in_one_day + 1 # update current_hour_in_one_day
        start_row_num = start_row_num + span_length # update span_length
        block_index = block_index + 1 # update block_index
        flag_data_time = False # update condition
        if (block_index + 1) % 24 == 0: # feedback which day dealed successfully
            print(" > [ Day " + str(day_num) + " / 365 ] finished !")
            day_num = day_num + 1
    # save the result to hard disk
    print(" > Saving to the hard disk now ...")
    result_wb.save(os.path.join(os.path.dirname(input_file), output_name)), print(" > Save finished ~")
    print(" > All data filds calculated finished ~"), print()

# Calculate the true length of a block
# The is function is aimed at the xlsx_file where some time-point lost
def getValidLength(begin_row_num, span, day):
    valid_data_length = 0 # initial length
    pre_minute = -1 # can not be 0
    for index in range(span):
        current_minute = sheet.cell(row = begin_row_num + index, column = 2).value.minute
        if current_minute <= 59 and current_minute > pre_minute: # base is 60
            valid_data_length = valid_data_length + 1
            pre_minute = current_minute
            continue
        break
    return valid_data_length # return the true length


# private function, used to calc a designated column's average number
def averageCol(sheet, begin_row_num, span, col_num):
    illegal_num = 0
    data_sum = 0
    for index in range(span):
        cell_value = sheet.cell(row = begin_row_num + index, column = col_num).value
        if str(cell_value)[0:1] == "-": # '-' means NaN
            illegal_num = illegal_num + 1
            continue
        data_sum = data_sum + cell_value # update data_sum
    if span == illegal_num:
        return 0
    else:
        return (data_sum / (span - illegal_num))


# Check the lost time-point and its range
def checkLostMinutes(row_num, ignore = 0):
    start_time_minute = -1
    count = 0
    range_index = 1
    for index in range(row_num - ignore):
        date_minute = sheet.cell(row = index + 1 + 2, column = 2).value.minute
        if date_minute < start_time_minute:  # the next number less than forward number
            delta_minute = date_minute + 60 - start_time_minute
        else:
            delta_minute = date_minute - start_time_minute
        if delta_minute > 1:
            count = count + (delta_minute - 1)
            print(" > Lost-minute No." + str(range_index) + " [ " + str(index + 1 + 1) + " - "\
             + str(index + 1 + 2) + " ]  lost numbers: " + str(delta_minute - 1))
            range_index += 1
        start_time_minute = date_minute
    print(" > Warning: There are " + str(count) + " time-point has lost !"), print()
    return count

def checkLostHours(row_num, ignore = 0):
    start_time_hour = -1
    count = 0
    range_index = 1
    for index in range(row_num - ignore):
        date_hour = sheet.cell(row = index + 1 + 2, column = 2).value.hour
        if (date_hour < start_time_hour) and (date_hour + 24 - start_time_hour >= 2) or (date_hour - start_time_hour >= 2):
            print(" > Lost-hour No." + str(range_index) + " [ " + str(index + 1 + 1) + " - "\
             + str(index + 1 + 2) + " ]  lost numbers: " + str(60))
            range_index += 1 # update range_index
            count = count + 1
        start_time_hour = date_hour
    print(" > Warning: There are " + str(count * 60) + " time-point has lost !"), print()
    return (count * 60)

def printLicense():
    print("------------------------------------------------------")
    print("Neverland_LY@163.com  2018-2019  All rights reserved.")
    print("------------------------------------------------------")

def alphaTest():
    pass

### PROGRAM ENTRANCE ###
if __name__== "__main__":
    printLicense()
    # open [.xlsx] file and check it's legality
    print(" > Loading the file ...")
    start_time = datetime.datetime.now()
    wb = openpyxl.load_workbook(input_file)
    end_time = datetime.datetime.now()
    print(" > Loading successful. Used %s seconds ~" % str((end_time - start_time).seconds))

    # check if the number of rows meets the requirements
    sheet = wb.active
    alphaTest()
    if (sheet.max_row - ignore_line) % base_span != 0:
        print(" > Warning: The file is not meets the requirements (means some time-point were lost) !")
        print()

    # check lost numbers
    lost_minute_count = checkLostMinutes(sheet.max_row, ignore_line), winsound.Beep(600,500)
    lost_hour_count = checkLostHours(sheet.max_row, ignore_line), winsound.Beep(600,500)
    print(" > Warning: Lost %s time-point totally !" % str(int(str(lost_minute_count[0])) +\
     int(str(lost_hour_count[0])))), print()
    sys.exit(1)

    # start dealing
    print(" > Start calculating the average of all data filds ~")
    calculateAllDataFildaverage(sheet.max_row, sheet.max_column, base_span, ignore_line)
    print(" > The program has exited ~"), print(), winsound.Beep(600,1000)
