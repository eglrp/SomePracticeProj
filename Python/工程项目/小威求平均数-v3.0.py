#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
# Author: Neverland_LY
# Date: 2018-11-26
# Version: v3.0
#
# Description:
#       1. This code is used fo calculate the average of the P, T and R value of the
#       meteorological data per day by month.
#       2. The scripts can only deal with the file ending with the suffix of .xlsx (except .xls).
#
# Usage:
#       1. Modify the input file path to your own path.
#       2. Do not modify the base_block number if unnecessary otherwise the program will crash and
#       result in some undefined behavior !

input_file = "E:\\xiaowei\\test.xlsx" # the file to be processed
output_name = "Result_LY.xlsx" # output name
base_span = 60 # produce an average every 60 lines
ignore_line = 2 # ignore the first 2 lines

#########################################################
#                                                       #
#   Do not change the following codes if unnecessary!   #
#                                                       #
#########################################################
import openpyxl
import os # Operator System library
import sys
import winsound
import datetime

# Calculate the p-R-T data's average
def calculateAllDataFildaverage(row_num, col_num, span = 60, ignore = 0):
    # define output file
    result_wb = openpyxl.Workbook()
    result_sheet = result_wb.active
    # read by line
    block_index = 0
    flag_data_time = False
    day_num = 0
    global sheet
    while block_index < int(row_num / span):
        date = "1997/1/1"
        for col_index in range(col_num):
            if flag_data_time is False:
                if col_index == 0:
                    date = sheet.cell(row = block_index * span + ignore + 1, column = col_index + 1).value
                    result_sheet.cell(row = block_index + 1, column = 1).value = date
                    continue
                if col_index == 1:
                    result_sheet.cell(row = block_index + 1, column = 2).value = str((block_index + 1) % 25)
                    flag_data_time = True
                    continue
            if col_index == 8 or col_index == 11:
                result_sheet.cell(row = block_index + 1, column = col_index + 1).value = "---"
                continue
            result_sheet.cell(row = block_index + 1, column = col_index + 1).value =\
                    averageCol(sheet, block_index * span + ignore + 1, col_index + 1, span)
        block_index = block_index + 1
        flag_data_time = False
        if block_index % 23 == 0:
            day_num = day_num + 1
            print(" > [ Day " + str(day_num) + " / 365 ] finished !")
            winsound.Beep(300,10)

    result_wb.save(os.path.join(os.path.dirname(input_file), output_name))
    print(" > calculateAllDataFildaverage finished !")


def averageCol(sheet, begin_row_num, col_num, line_span):
    illegal_num = 0
    data_sum = 0
    for index in range(line_span):
        cell_value = sheet.cell(row = begin_row_num + index, column = col_num).value
        if cell_value == "---":
            illegal_num = illegal_num + 1
            continue
        data_sum = data_sum + cell_value
    return (data_sum / (line_span - illegal_num))

def checkLossTime(row_num, ignore = 0):
    start_time_minute = -1
    count = 0
    range_index = 1
    for index in range(row_num - ignore):
        date = sheet.cell(row = index + 1 + 2, column = 2).value
        if date.minute < start_time_minute:  # the next number less than forward number
            delta_minute = date.minute + 60 - start_time_minute
        else:
            delta_minute = date.minute - start_time_minute
        if delta_minute > 1:
            count = count + (delta_minute - 1)
            print(" > Lost-Block No." + str(range_index) + " [ " + str(index + 1 + 1) + " - " + str(index + 1 + 2) + " ]  lost numbers: " + str(delta_minute - 1))
            range_index += 1
        start_time_minute = date.minute
    print(" >>> " + str(count) + " data lost !")


### PROGRAM ENTRANCE ###
if __name__== "__main__":
    # open [.xlsx] file and check it's legality
    print(" > Loading the file ...")
    wb = openpyxl.load_workbook(input_file)
    print(" > Loading successful ...")
    sheet = wb.active
    if (sheet.max_row - ignore_line) % base_span != 0:
        print(" > The file is not legal !")
        # sys.exit(1)

    # check lost numbers
    winsound.Beep(600,1000)
    checkLossTime(sheet.max_row, 2)
    winsound.Beep(600,1000)
    sys.exit(1)

    # start dealing
    print(" > Start running ~"), winsound.Beep(600,1000)
    calculateAllDataFildaverage(sheet.max_row, sheet.max_column, base_span, ignore_line)
    print(" > The program has exited ~"), winsound.Beep(600,1000)