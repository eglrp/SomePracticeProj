import openpyxl

# 打开文件
wb = openpyxl.load_workbook('C:/Users/Neverland_LY/Desktop/xw.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

# 定义输出文件
result_wb = openpyxl.Workbook()
result_sheet = result_wb.get_active_sheet()
list_col = ['A', 'B', 'C', 'D', 'E', 'F', 'G','H', 'E', 'J', 'K']

# 设置间隔行数
scan_raw = 8

for i in range(sheet.max_row - scan_raw + 1):
    for j in range(scan_raw):
        result_sheet[list_col[j] + str(i + 1)] = \
        sheet['A' + str(i + 1 + j)].value

# 保存文件
result_wb.save('每隔' + str(scan_raw) + '行.xlsx')