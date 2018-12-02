#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
# Author: Neverland_LY
# Date: 2018-11-21
# Version: v1.0
#
# Description:
#       1. This code is used fo calculate the average of the P, T and R value of the
#       meteorological data per day by month.
#       2. The scripts can only deal with the file ending with the suffix of .xlsx (except .xls).
#
# Usage:
#       1. Modify the input file path to your own path.
#       2. DON'T modify the base_block number if unnecessary otherwise the program will crash and
#       result in some undefined behavior !

import openpyxl
import os # Operator System library

# define input path (the content of path can contains both folders and files)
input_file_path = "C:\\Users\\Neverland_LY\\Desktop\\xiaowei" # do not add \\
# define output path (the path can be any folder)
output_path = "E:\\xiaowei"  # do not add \\ in the end of path and the output path must have been created firstly by user
#define the block number
base_block = 120 # WARNING #

# define the counter
count = 1
# get all files that the names ending with the suffix of [.xlsx]
for root_path, sub_directory_list, un_sub_directory_list in os.walk(input_file_path):
    for file in un_sub_directory_list:
        # open [.xlsx] file
        wb = openpyxl.load_workbook(root_path + "\\" + file)
        # get active sheet
        active_sheet = wb.active
        # define output file
        result_wb = openpyxl.Workbook()
        result_sheet = result_wb.active
        # merge 120 lines to 1 line
        for block in range(int(active_sheet.max_row / base_block)): # 0 - 24
            # pick the year-month-day
            block_year = active_sheet['A1'].value
            block_month = active_sheet['B1'].value
            block_day = active_sheet['C1'].value
            # define the sum of P T R
            sum_P_COLG = 0.000; sum_T_COLH = 0.000; sum_R_COLI = 0.000;
            for sub_line in range(base_block): # 0 - 120
                # confirm the row number
                row_num = block * base_block + sub_line + 1
                # loop for the summation
                sum_P_COLG += active_sheet['G' + str(row_num)].value
                sum_T_COLH += active_sheet['H' + str(row_num)].value
                sum_R_COLI += active_sheet['I' + str(row_num)].value
            # calc average of P T R
            sum_P_COLG /= base_block; sum_T_COLH /= base_block; sum_R_COLI /= base_block;
            # optput to the new [.xlsx] file
            # fill the year-month-day-hour
            result_sheet['A' + str(block + 1)] = block_year
            result_sheet['B' + str(block + 1)] = block_month
            result_sheet['C' + str(block + 1)] = block_day
            result_sheet['D' + str(block + 1)] = block
            # fill the P-T-R
            result_sheet['G' + str(block + 1)] = sum_P_COLG
            result_sheet['H' + str(block + 1)] = sum_T_COLH
            result_sheet['I' + str(block + 1)] = sum_R_COLI
        # save to the hard-disk
        result_wb.save(output_path + '\\' + file[:-5] + '-LY.xlsx')
        print(" [ No." + str(count) + " ] "+ file + " operate successful~")
        count += 1
        # break
    # break
print("All tasks finished~")
print("The program has exited~")