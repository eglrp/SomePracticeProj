#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
# Author: Neverland_LY
# Date: 2018-11-22
# Version: v2.0
#
# Description:
#       1. This code is used fo calculate the average of the P, T and R value of the
#       meteorological data per day by month.
#       2. The scripts can only deal with the file ending with the suffix of .xlsx (except .xls).
#
# Usage:
#       1. Modify the input file path to your own path.
#       2. Do not modify the base_block number if unnecessary otherwise the program will crash and
#       result in some undefined behavior !


# define input path (the content of path can contains both folders and files)
global_input_path = "C:\\Users\\Neverland_LY\\Desktop\\xiaowei" # do not add \\ in th end
# define output path (the path can be any folder)
global_output_path = "E:\\xiaowei"  # do not add \\ in th end and the output path must have been created firstly by user
# define output name
global_output_name = "Result-LY.xlsx" # the file must ending with [.xlsx]
#define the block number
global_base_block = 120 # WARNING #

#########################################################
#                                                       #
#   Do not change the following codes if unnecessary!   #
#                                                       #
#########################################################
import openpyxl
import os # Operator System library

# Calculate the p-R-T data's average
def calc_PRT_average(input_path, output_path, output_name, base_block = 120):
    # define the counter
    count = 0
    # define output file
    result_wb = openpyxl.Workbook()
    result_sheet = result_wb.active
    # get all files that the names ending with the suffix of [.xlsx]
    for root_path, sub_directory_list, un_sub_directory_list in os.walk(input_path):
        for file in un_sub_directory_list:
            # open [.xlsx] file
            wb = openpyxl.load_workbook(root_path + "\\" + file)
            # get active sheet
            active_sheet = wb.active
            # merge 120 lines to 1 line
            base_line = int(active_sheet.max_row / base_block)
            for block in range(base_line): # [0, 24)
                # pick the year-month-day
                block_year = active_sheet['A1'].value
                block_month = active_sheet['B1'].value
                block_day = active_sheet['C1'].value
                # define the sum of P T R
                sum_P_COLG = 0.000; sum_T_COLH = 0.000; sum_R_COLI = 0.000;
                for sub_line in range(base_block): # [0, 120)
                    # confirm the row number
                    row_num = block * base_block + sub_line + 1
                    # loop for the summation
                    sum_P_COLG += active_sheet['G' + str(row_num)].value
                    sum_T_COLH += active_sheet['H' + str(row_num)].value
                    sum_R_COLI += active_sheet['I' + str(row_num)].value
                # calc average of P T R
                sum_P_COLG /= base_block; sum_T_COLH /= base_block; sum_R_COLI /= base_block;
                # optput to the new [.xlsx] file
                # fill the year-month-day-hour
                result_sheet['A' + str(block + 1 + count * base_line)] = block_year
                result_sheet['B' + str(block + 1 + count * base_line)] = block_month
                result_sheet['C' + str(block + 1 + count * base_line)] = block_day
                result_sheet['D' + str(block + 1 + count * base_line)] = block
                # fill the P-T-R
                result_sheet['G' + str(block + 1 + count * base_line)] = sum_P_COLG
                result_sheet['H' + str(block + 1 + count * base_line)] = sum_T_COLH
                result_sheet['I' + str(block + 1 + count * base_line)] = sum_R_COLI
            # feedback to users
            print("     [ No." + str(count + 1) + " ] "+ file + " operate successful ~")
            count += 1
    # save to the hard-disk
    if not os.path.exists(output_path):
        os.makedirs(output_path)
    print(" > Start merging all datas ~")
    result_wb.save(output_path + '\\' + output_name)
    print(" > Merging finished ~")
    return

### PROGRAM ENTRANCE ###
if __name__== "__main__":
    print(" > Start execting ~")
    calc_PRT_average(global_input_path, global_output_path, global_output_name, global_base_block)
    print(" > The program has exited ~")